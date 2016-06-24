import openpyxl
import time
import datetime
from datetime import timedelta
import time
import smtplib
import re
from openpyxl.drawing.image import Image
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email import encoders
from constants import *

fromaddr = "appdynamicshackthon@gmail.com"

#TODO THIS IS VERY UGLY, WILL CHANGE IF TIME
def parse_email(mailtolink):
	regex = '<mailto.*\|'
	m = re.search(regex, mailtolink)
	res = m.group(0)
	return res[8:-1]

def closest_sunday(d):
	days_of_week = ['sunday','monday','tuesday','wednesday', 'thursday','friday','saturday']
	today = d.isoweekday()
	if(today > 4):
		delta_day = 7-today
	else:
		delta_day = 0 - d.isoweekday()
	return d + timedelta(days=delta_day)

def add_signature_image(sheet):
	xfile = openpyxl.load_workbook(filepath)
	sheet = xfile.get_sheet_by_name('Time Card')
	img = Image('signature_scan.gif')
	img.drawing.width = 100
	img.drawing.height = 50
	img.anchor(sheet.cell('E35'))
	sheet.add_image(img)

def generate_specific(fullname, date, job, mon, tue, wed, thur, fri):
	"""
	the date should be in the format mm-dd-yyyy
	"""
	xfile = openpyxl.load_workbook(filepath)
	sheet = xfile.get_sheet_by_name('Time Card')
	sheet['B19'] = job
	sheet['C8'] = fullname
	sheet['C19'] = mon
	sheet['E19'] = tue
	sheet['G19'] = wed
	sheet['I19'] = thur
	sheet['K19'] = fri
	sheet['O8'] = date.strftime('%m/%d/%Y')
	sheet['B35'] = fullname
	sheet['H35'] = datetime.date.today().strftime('%m/%d/%Y')
	new_file_name = './temp/' + fullname + '_Timesheet' + '_' + date.strftime('%m%d%Y') + '_' + '.xlsx'
	xfile.save(new_file_name)
	return new_file_name

def send_email(name, email,d, filename, filepath, manager):
	toaddr = parse_email(email)
	msg = MIMEMultipart()
	msg['From'] = fromaddr
	msg['To'] = toaddr
	last_week = d - timedelta(days=6)
	date = last_week.strftime('%m/%d') + " - " + d.strftime('%m/%d')
	msg['Subject'] = "Timesheet: " + date

	toperson = manager
	body = "Dear " + toperson + ",\n" + "\nAttached is the timesheet for your approval.\n" + "\nBest,\n" + name

	msg.attach(MIMEText(body, 'plain'))
	attachment = open(filepath, "rb")

	part = MIMEBase('application', 'octet-stream')
	part.set_payload((attachment).read())
	encoders.encode_base64(part)
	part.add_header('Content-Disposition', "attachment; filename= %s" % filename)
	msg.attach(part)

	server = smtplib.SMTP('smtp.gmail.com', 587)
	server.starttls()
	server.login(fromaddr, "slackthon") #put password here
	text = msg.as_string()
	server.sendmail(fromaddr, toaddr, text)
	server.quit()


